
"""
Convert ONS LSOA broad-age population estimate Excel workbooks to a clean yearly Parquet file.

This parser is tailored to ONS broad-age LSOA workbooks whose population sheets
look like:
    row 1-3: title/notes
    row 4:   headers, including "LSOA 2021 Code", "LSOA 2021 Name", "Total"
    row 5+:  data

Output schema:
    lsoa_code
    lsoa_name
    population_year
    population
    source_file
    source_sheet

Run from repo root:
    python -m src.cbl20.population_offset
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook


DEFAULT_OLD_XLSX = Path("data/external/lsoa_population_mid2011_mid2022.xlsx")
DEFAULT_REVISED_XLSX = Path("data/external/lsoa_population_mid2022revised_mid2024.xlsx")
DEFAULT_OUTPUT = Path("data/processed/lsoa_population_yearly.parquet")


def _extract_year_from_sheet_name(sheet_name: str) -> int | None:
    match = re.search(r"Mid-(\d{4})\s+LSOA", sheet_name)
    if not match:
        return None
    return int(match.group(1))


def _clean_lsoa_code(value) -> str | None:
    if value is None:
        return None
    code = str(value).strip()
    if re.match(r"^[EW]\d{8}$", code):
        return code
    return None


def _parse_population_sheet_openpyxl(xlsx_path: Path, sheet_name: str, year: int) -> pd.DataFrame:
    """
    Fast parser for one ONS population sheet using openpyxl read-only mode.

    We avoid pandas.read_excel here because these workbooks are large and have
    wide sheets. The needed columns are only:
      - LSOA 2021 Code
      - LSOA 2021 Name
      - Total
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]

        # Header is row 4 in the ONS workbook.
        header_row = next(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        headers = [str(x).strip() if x is not None else "" for x in header_row]

        required = {
            "LSOA 2021 Code": None,
            "LSOA 2021 Name": None,
            "Total": None,
        }

        for idx, header in enumerate(headers):
            if header in required:
                required[header] = idx

        missing = [name for name, idx in required.items() if idx is None]
        if missing:
            raise ValueError(
                f"Sheet {sheet_name!r} missing required columns {missing}. "
                f"Headers found: {headers[:20]}"
            )

        code_idx = required["LSOA 2021 Code"]
        name_idx = required["LSOA 2021 Name"]
        total_idx = required["Total"]

        rows = []
        max_idx = max(code_idx, name_idx, total_idx)

        for row in ws.iter_rows(min_row=5, values_only=True):
            if row is None or len(row) <= max_idx:
                continue

            lsoa_code = _clean_lsoa_code(row[code_idx])
            if lsoa_code is None:
                continue

            lsoa_name = "" if row[name_idx] is None else str(row[name_idx]).strip()

            try:
                population = float(row[total_idx])
            except (TypeError, ValueError):
                continue

            if not np.isfinite(population) or population <= 0:
                continue

            rows.append(
                {
                    "lsoa_code": lsoa_code,
                    "lsoa_name": lsoa_name,
                    "population_year": int(year),
                    "population": int(round(population)),
                    "source_file": Path(xlsx_path).name,
                    "source_sheet": sheet_name,
                }
            )

        if not rows:
            raise ValueError(f"Sheet {sheet_name!r} produced no valid LSOA population rows.")

        return pd.DataFrame(rows)

    finally:
        wb.close()


def parse_ons_lsoa_population_excel(xlsx_path: Path) -> pd.DataFrame:
    """
    Parse all 'Mid-YYYY LSOA 2021' sheets from one ONS workbook.
    """
    xlsx_path = Path(xlsx_path)

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Population workbook not found: {xlsx_path}")

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet_names = list(wb.sheetnames)
    finally:
        wb.close()

    parsed = []
    errors = []

    target_sheets = [(sheet, _extract_year_from_sheet_name(sheet)) for sheet in sheet_names]
    target_sheets = [(sheet, year) for sheet, year in target_sheets if year is not None]

    print(f"Parsing {xlsx_path.name}: {len(target_sheets)} yearly LSOA sheets found")

    for sheet_name, year in target_sheets:
        try:
            print(f"  - parsing {sheet_name} ...", flush=True)
            parsed.append(_parse_population_sheet_openpyxl(xlsx_path, sheet_name, year))
        except Exception as exc:
            errors.append(f"{sheet_name}: {exc}")

    if not parsed:
        msg = (
            f"Could not parse any population sheet from {xlsx_path}.\n"
            f"Sheets found: {sheet_names}\n"
        )
        if errors:
            msg += "Parsing errors:\n" + "\n".join(errors[:10])
        raise ValueError(msg)

    population = pd.concat(parsed, ignore_index=True)

    if population.empty:
        raise ValueError(f"Parsed workbook but found no valid rows: {xlsx_path}")

    return population


def combine_population_editions(
    old_2011_2022_path: Path,
    revised_2022_2024_path: Path,
) -> pd.DataFrame:
    """
    Combine editions.

    Rule:
    - use 2011-2021 from old edition
    - use 2022-2024 from revised edition
    """
    old = parse_ons_lsoa_population_excel(old_2011_2022_path)
    revised = parse_ons_lsoa_population_excel(revised_2022_2024_path)

    old_keep = old[old["population_year"] <= 2021].copy()
    revised_keep = revised[revised["population_year"] >= 2022].copy()

    population = pd.concat([old_keep, revised_keep], ignore_index=True)

    population = (
        population
        .sort_values(["lsoa_code", "population_year", "source_file"])
        .drop_duplicates(["lsoa_code", "population_year"], keep="last")
        .sort_values(["lsoa_code", "population_year"])
        .reset_index(drop=True)
    )

    return population


def validate_population_table(population: pd.DataFrame) -> dict:
    diagnostics = {
        "n_rows": int(len(population)),
        "n_lsoas": int(population["lsoa_code"].nunique()),
        "min_year": int(population["population_year"].min()),
        "max_year": int(population["population_year"].max()),
        "min_population": int(population["population"].min()),
        "max_population": int(population["population"].max()),
        "missing_population_rows": int(population["population"].isna().sum()),
        "duplicate_lsoa_year_rows": int(population.duplicated(["lsoa_code", "population_year"]).sum()),
    }

    if diagnostics["duplicate_lsoa_year_rows"] > 0:
        raise ValueError(f"Found duplicated LSOA-year rows: {diagnostics['duplicate_lsoa_year_rows']}")
    if diagnostics["missing_population_rows"] > 0:
        raise ValueError(f"Found missing population rows: {diagnostics['missing_population_rows']}")
    if diagnostics["min_population"] <= 0:
        raise ValueError("Population contains non-positive values.")

    return diagnostics


def build_population_parquet(
    old_2011_2022_path: Path,
    revised_2022_2024_path: Path,
    output_path: Path,
) -> tuple[pd.DataFrame, dict]:
    population = combine_population_editions(old_2011_2022_path, revised_2022_2024_path)
    diagnostics = validate_population_table(population)

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    population.to_parquet(output_path, index=False)

    return population, diagnostics


def load_population_parquet(parquet_path: Path) -> pd.DataFrame:
    return pd.read_parquet(parquet_path)


def add_population_offset_to_panel(
    panel: pd.DataFrame,
    population: pd.DataFrame,
    *,
    month_col: str = "month",
    lsoa_col: str = "lsoa_code",
    latest_year_fallback: bool = True,
    fill_missing: bool = True,
) -> pd.DataFrame:
    """
    Add population and log_offset to an LSOA-month panel.

    For 2025/2026 rows, if latest_year_fallback=True, it uses the latest
    available population estimate year, e.g. 2024.
    """
    out = panel.copy()
    pop = population.copy()

    pop["lsoa_code"] = pop["lsoa_code"].astype(str).str.strip()
    out[lsoa_col] = out[lsoa_col].astype(str).str.strip()

    out["population_year_requested"] = pd.to_datetime(out[month_col]).dt.year
    latest_year = int(pop["population_year"].max())

    if latest_year_fallback:
        out["population_year_used"] = out["population_year_requested"].clip(upper=latest_year)
    else:
        out["population_year_used"] = out["population_year_requested"]

    out = out.merge(
        pop[["lsoa_code", "population_year", "population"]],
        left_on=[lsoa_col, "population_year_used"],
        right_on=["lsoa_code", "population_year"],
        how="left",
        suffixes=("", "_pop"),
    )

    if "lsoa_code_pop" in out.columns:
        out = out.drop(columns=["lsoa_code_pop"])
    if "population_year" in out.columns:
        out = out.drop(columns=["population_year"])

    missing_rate = out["population"].isna().mean()

    if missing_rate > 0 and not fill_missing:
        raise ValueError(f"Missing population for {missing_rate:.2%} of panel rows.")

    if fill_missing and missing_rate > 0:
        out["population"] = out["population"].fillna(out["population"].median())

    if (out["population"] <= 0).any():
        raise ValueError("Population values must be positive before taking log offset.")

    out["log_offset"] = np.log(out["population"])

    return out


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert ONS LSOA population estimate Excel files into yearly Parquet."
    )
    parser.add_argument("--old-xlsx", type=Path, default=None)
    parser.add_argument("--revised-xlsx", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    repo_root = Path.cwd()

    old_xlsx = args.old_xlsx or repo_root / DEFAULT_OLD_XLSX
    revised_xlsx = args.revised_xlsx or repo_root / DEFAULT_REVISED_XLSX
    output = args.output or repo_root / DEFAULT_OUTPUT

    print("Converting ONS LSOA population estimates...")
    print(f"Old edition:     {old_xlsx}")
    print(f"Revised edition: {revised_xlsx}")
    print(f"Output:          {output}")

    population, diagnostics = build_population_parquet(old_xlsx, revised_xlsx, output)

    print("\nDone.")
    print("Diagnostics:")
    for key, value in diagnostics.items():
        print(f"  {key}: {value}")

    print("\nPreview:")
    print(population.head().to_string(index=False))


if __name__ == "__main__":
    main()
